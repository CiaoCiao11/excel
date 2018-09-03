from openpyxl import Workbook
wb = Workbook()
ws1=wb.create_sheet('Mysheet')#默認插入到最後
ws2=wb.create_sheet('Mysheet1',0)#指定插入到第一個位置

# grab the active worksheet
ws = wb.active
ws.title='New Title'

ws3=wb['New Title']

# Data can be assigned directly to cells
ws['A1'] = 'chinese'
ws['B1'] = 'english'
ws['C1'] = 'math'
ws['D1'] = 'physics'

# Rows can also be appended
ws.append([1, 2, 3])

# Python types will automatically be converted
import datetime
ws['A2'] = datetime.datetime.now()

# Save the file
wb.save("score.xlsx")