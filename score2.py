from openpyxl import Workbook
from openpyxl import load_workbook
#  如果是我的應用展式，我是先讀取舊的excel檔，然後同時開一個新的Workbook(就是excel檔)，
# 所以檔案開頭是這樣：(wb代表workbook, ws代表worksheet，也是excel檔裡的分頁，一個workbook可以有很多個worksheet)
wb = load_workbook('score.xlsx')
ws = wb.active
wb_new = Workbook()
ws_new = wb_new.active



# 然後我開始把舊的excel檔的內容 一格一格看我要哪些，貼到新的worksheet (也就是ws_new)。
ws_new['B1'].value = ws['A1'].value
ws_new['C1'].value = ws['B1'].value
ws_new['D1'].value = ws['C1'].value

wb_new.save('newscore.xlsx')